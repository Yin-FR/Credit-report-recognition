#!/usr/bin/env python3
# -*-coding: utf-8-*-
# @File: result_process.py
# @Author: 檀寅
# @Time: 2020年08月19日09:48
# @说明: 未用

import openpyxl

def result_process(index, xlsxPath):
    if index == '1':
        workbook = openpyxl.load_workbook(xlsxPath)
        position_r_0, position_c_0 = (0, 0)
        position_r_1, position_c_1 = (0, 0)
        ws_0, ws_1 = (None, None)
        for ws in workbook.worksheets:
            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if ws.cell(row=i, column=j).value == '笔数':
                        position_r_0 = i
                        position_c_0 = j
                        ws_0 = ws

        for ws in workbook.worksheets:
            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if ws.cell(row=i, column=j).value == '发卡法人机构数':
                        position_r_1 = i
                        position_c_1 = j
                        ws_1 = ws

        if (position_c_0 and position_r_0):
            ws_0.cell(position_r_0, position_c_0 + 1).value = '月份数'
            ws_0.cell(position_r_0, position_c_0 + 2).value = '单月最高逾期总额'
            ws_0.cell(position_r_0, position_c_0 + 3).value = '最长逾期月数'
            ws_0.cell(position_r_0, position_c_0 + 4).value = '账户数'
            ws_0.cell(position_r_0, position_c_0 + 5).value = '月份数'
            ws_0.cell(position_r_0, position_c_0 + 6).value = '单月最高逾期总额'
            ws_0.cell(position_r_0, position_c_0 + 7).value = '最长逾期月数'
            ws_0.cell(position_r_0, position_c_0 + 8).value = '账户数'
            ws_0.cell(position_r_0, position_c_0 + 9).value = '月份数'
            ws_0.cell(position_r_0, position_c_0 + 10).value = '单月最高透支余额'
            ws_0.cell(position_r_0, position_c_0 + 11).value = '最长透支月数'

        if (position_r_1 and position_c_1):
            ws_1.cell(position_r_1, position_c_1 + 1).value = '发卡机构数'
            ws_1.cell(position_r_1, position_c_1 + 2).value = '账户数'
            ws_1.cell(position_r_1, position_c_1 + 3).value = '授信总额'
            ws_1.cell(position_r_1, position_c_1 + 4).value = '单家行最高授信额'
            ws_1.cell(position_r_1, position_c_1 + 5).value = '单家行最低授信额'
            ws_1.cell(position_r_1, position_c_1 + 6).value = '已用额度'
            ws_1.cell(position_r_1, position_c_1 + 7).value = '最近6个月平均使用额度'


        workbook.save(xlsxPath)

    else:
        pass